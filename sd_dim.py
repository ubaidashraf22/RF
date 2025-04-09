import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import pandas as pd
import numpy as np
import math
import re
import warnings
warnings.filterwarnings('ignore')

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

##############################################
# Erlang B function (unchanged)
##############################################
def erlangB(A, N):
    """Computes Erlang-B blocking probability for offered load A and N channels."""
    if A <= 0:
        return 0  
    if N <= 0:
        return 1
    
    invB = 1.0
    for i in range(1, N + 1):
        invB = 1.0 + (i / A) * invB
    return 1.0 / invB

##############################################
# The main GUI class
##############################################
class SDCCHDimensioningGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("SDCCH Dimensioning Tool")

        # File paths
        self.sdcch_csv_path = None
        self.trxchan_txt_path = None

        # DataFrames
        self.df_sd = None
        self.df_trxchan = None

        # For storing date selection
        self.date_vars = {}  # {date: BooleanVar()}
        self.dates_window = None

        # Create top-level frames/panels
        self.create_widgets()

    def create_widgets(self):
        """Create and place widgets on the main window."""
        # Instructions at the top
        instructions_text = (
            "Instructions:\n"
            "1) Click 'Load SDCCH CSV' to upload the SDCCH hourly stats (in background).\n"
            "2) Click 'Load TRXCHAN TXT' to upload the TRXCHAN configuration.\n"
            "3) You can deselect any abnormal dates after loading SDCCH.\n"
            "4) Enter how many top busy days to consider for dimensioning.\n"
            "5) Finally, click 'Process Data'."
        )
        self.instructions_label = tk.Label(self.master, text=instructions_text, justify="left")
        self.instructions_label.pack(pady=(10,5))

        # Frame for SDCCH
        self.sdcch_frame = tk.Frame(self.master, bd=2, relief="groove")
        self.sdcch_frame.pack(fill="x", padx=5, pady=5)

        self.sdcch_button = tk.Button(self.sdcch_frame, text="Load SDCCH CSV", command=self.load_sdcch_in_background)
        self.sdcch_button.pack(side="left", padx=5, pady=5)

        # Indeterminate progress bar for loading SDCCH
        self.sdcch_progress = ttk.Progressbar(self.sdcch_frame, orient='horizontal', mode='determinate', length=200)
        self.sdcch_progress.pack(side="left", padx=5, pady=5)

        self.sdcch_status = tk.Label(self.sdcch_frame, text="No file loaded.")
        self.sdcch_status.pack(side="left", padx=5, pady=5)

        # Frame for TRXCHAN
        self.trxchan_frame = tk.Frame(self.master, bd=2, relief="groove")
        self.trxchan_frame.pack(fill="x", padx=5, pady=5)

        self.trxchan_button = tk.Button(self.trxchan_frame, text="Load TRXCHAN TXT", command=self.load_trxchan)
        self.trxchan_button.pack(side="left", padx=5, pady=5)

        self.trxchan_status = tk.Label(self.trxchan_frame, text="No file loaded.")
        self.trxchan_status.pack(side="left", padx=5, pady=5)

        # Frame for top-busy-days input
        self.top_days_frame = tk.Frame(self.master, bd=2, relief="groove")
        self.top_days_frame.pack(fill="x", padx=5, pady=5)

        self.top_days_label = tk.Label(self.top_days_frame, text="Number of top busy days to consider:")
        self.top_days_label.pack(side="left", padx=5, pady=5)
        
        self.top_days_var = tk.StringVar(value="5")  # default to 5
        self.top_days_entry = tk.Entry(self.top_days_frame, textvariable=self.top_days_var, width=5)
        self.top_days_entry.pack(side="left", padx=5, pady=5)

        # Comment: "select SDCCH busy hour "" out of total"
        # You could add a label or comment here to reflect that logic if needed:
        self.comment_label = tk.Label(self.top_days_frame, text='(Select SDCCH busy hours out of total...)', fg='blue')
        self.comment_label.pack(side="left", padx=5)

        # Process Data button + progress bar
        self.process_frame = tk.Frame(self.master, bd=2, relief="groove")
        self.process_frame.pack(fill="x", padx=5, pady=5)

        self.process_button = tk.Button(self.process_frame, text="Process Data", command=self.process_data_in_background, state=tk.DISABLED)
        self.process_button.pack(side="left", padx=5, pady=5)

        self.process_progress = ttk.Progressbar(self.process_frame, orient='horizontal', mode='determinate', length=200)
        self.process_progress.pack(side="left", padx=5, pady=5)

        self.process_status = tk.Label(self.process_frame, text="")
        self.process_status.pack(side="left", padx=5, pady=5)

    ##################################
    # 1. LOAD SDCCH CSV (background)
    ##################################
    def load_sdcch_in_background(self):
        """Spawn a thread to load SDCCH so the GUI doesn't freeze."""
        thread = threading.Thread(target=self.load_sdcch, daemon=True)
        thread.start()

    def load_sdcch(self):
        """Actual code to load the SDCCH CSV in a separate thread."""
        # Reset progress bar
        self.sdcch_progress.config(mode='indeterminate')
        self.sdcch_progress.start(10)

        file_path = filedialog.askopenfilename(
            title="Select SDCCH CSV File",
            filetypes=(("CSV Files", "*.csv"), ("All Files", "*.*"))
        )
        if not file_path:
            self.sdcch_progress.stop()
            self.sdcch_progress.config(mode='determinate', value=0)
            return

        self.sdcch_csv_path = file_path
        try:
            # Read CSV (entire file at once—this is just a demo)
            # Large files might need chunk processing.
            df_sdcch = pd.read_csv(self.sdcch_csv_path, skiprows=6, skipfooter=1, engine='python')
            self.df_sd = df_sdcch.copy()

            # Convert 'Time' to datetime
            self.df_sd['Time'] = pd.to_datetime(self.df_sd['Time'], errors='coerce')
            self.df_sd['Date'] = self.df_sd['Time'].dt.date

            unique_dates = sorted(self.df_sd['Date'].dropna().unique())
            msg = f"Loaded. Found {len(unique_dates)} unique dates."
            self.sdcch_status.config(text=msg)

            # Now show date selection window (on the main thread)
            self.master.after(0, self.open_dates_window, unique_dates)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to read SDCCH CSV:\n{e}")
        finally:
            # Stop progress bar
            self.sdcch_progress.stop()
            self.sdcch_progress.config(mode='determinate', value=100)
            # Possibly enable process button if TRXCHAN is also loaded
            if self.df_trxchan is not None:
                self.enable_process_button()

    ##################################
    # 2. SHOW A WINDOW FOR DATE SELECTION
    ##################################
    def open_dates_window(self, date_list):
        """
        Creates a Toplevel window with checkboxes for each date.
        User can uncheck to exclude them as abnormal dates.
        """
        if self.dates_window is not None and tk.Toplevel.winfo_exists(self.dates_window):
            # If there's already an open date window, destroy it first
            self.dates_window.destroy()

        self.dates_window = tk.Toplevel(self.master)
        self.dates_window.title("Select Dates to INCLUDE")

        info_label = tk.Label(self.dates_window, text="Uncheck any abnormal date(s) to EXCLUDE from processing:")
        info_label.pack(padx=5, pady=5)

        # Create a scrolling frame if there are many dates
        frame_canvas = tk.Frame(self.dates_window)
        frame_canvas.pack(fill="both", expand=True)

        canvas = tk.Canvas(frame_canvas)
        scrollbar = tk.Scrollbar(frame_canvas, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Populate checkboxes
        self.date_vars.clear()
        for dt in date_list:
            var = tk.BooleanVar(value=True)  # default: date is selected
            chk = tk.Checkbutton(scrollable_frame, text=str(dt), variable=var)
            chk.pack(anchor="w", padx=5, pady=2)
            self.date_vars[dt] = var

        # A close button
        close_button = tk.Button(self.dates_window, text="Close", command=self.dates_window.destroy)
        close_button.pack(pady=5)

    ##################################
    # 3. LOAD TRXCHAN TXT
    ##################################
    def load_trxchan(self):
        file_path = filedialog.askopenfilename(
            title="Select TRXCHAN TXT File",
            filetypes=(("TXT Files", "*.txt"), ("All Files", "*.*"))
        )
        if not file_path:
            return

        self.trxchan_txt_path = file_path
        try:
            self.df_trxchan = pd.read_csv(self.trxchan_txt_path, header=1, encoding='unicode_escape')
            self.trxchan_status.config(text="TRXCHAN loaded successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read TRXCHAN TXT:\n{e}")
            return

        if self.df_sd is not None:
            self.enable_process_button()

    def enable_process_button(self):
        self.process_button.config(state=tk.NORMAL)

    ##################################
    # 4. PROCESS DATA (in background)
    ##################################
    def process_data_in_background(self):
        """Spawn a thread to process data so the GUI doesn't freeze."""
        thread = threading.Thread(target=self.process_data, daemon=True)
        thread.start()

    def process_data(self):
        """Main dimensioning logic, run in a separate thread."""
        # Start progress bar
        self.process_progress.config(mode='indeterminate')
        self.process_progress.start(10)
        self.process_status.config(text="Processing...")

        try:
            if self.df_sd is None or self.df_trxchan is None:
                messagebox.showwarning("Warning", "Please load both SDCCH CSV and TRXCHAN TXT.")
                return

            # Copy the data
            df_sd = self.df_sd.copy()

            # Exclude any user-deselected dates
            for dt, var in self.date_vars.items():
                if not var.get():
                    # If user unchecked it, remove from dataset
                    df_sd = df_sd[df_sd['Date'] != dt]

            # Let the user specify the top N busy days
            try:
                top_n = int(self.top_days_var.get())
            except ValueError:
                top_n = 5  # fallback if user typed something invalid

            # Convert the 'Total SDCCH Traffic (Erl)_South' column to numeric
            df_sd['Total SDCCH Traffic (Erl)_South'] = pd.to_numeric(df_sd['Total SDCCH Traffic (Erl)_South'], errors='coerce')

            # Group by cell+date, find max daily traffic
            df_max_sd = df_sd.groupby(['Cell CI', 'Date'])['Total SDCCH Traffic (Erl)_South'].max().reset_index()

            # Take the top N largest traffic days, then average them
            df_avg_topN = (
                df_max_sd.groupby('Cell CI')
                .apply(lambda x: x.nlargest(top_n, 'Total SDCCH Traffic (Erl)_South')['Total SDCCH Traffic (Erl)_South'].mean())
                .reset_index()
            )
            df_avg_topN.columns = ['Cell CI', 'Max N Days Avg SDCCH Traffic (Erl)_South']

            # We'll also track the last date's "SDCCH8" for demonstration,
            # or we can do something else if needed:
            last_date = df_sd['Date'].max()
            df_last_date = df_sd[df_sd['Date'] == last_date].copy()
            df_last_date['SDCCH8'] = np.ceil(df_last_date['Initially Configured SDCCH_South'] / 8)

            # Merge
            df_avg_topN = pd.merge(df_last_date, df_avg_topN, on='Cell CI', how='left')

            # Compute required channels using blocking probability
            blocking_probability = 0.001
            required_channels = []
            for traffic in df_avg_topN['Max N Days Avg SDCCH Traffic (Erl)_South']:
                if pd.isna(traffic):
                    # If no data, assume 0 or skip
                    required_channels.append(0)
                    continue
                channels = 1
                p_blocking = 1
                while p_blocking > blocking_probability:
                    p_blocking = erlangB(traffic, channels)
                    if p_blocking > blocking_probability:
                        channels += 1
                required_channels.append(channels)

            df_avg_topN['Required Channels'] = required_channels
            df_avg_topN['Required SDCCH8'] = np.ceil(df_avg_topN['Required Channels'] / 8)
            df_avg_topN['Diff'] = df_avg_topN['Required SDCCH8'] - df_avg_topN['SDCCH8']

            # Comments: negative => convert TCH -> SDCCH, positive => we need more SDCCH
            df_avg_topN['Comments'] = [
                f"{int(-diff)} TCHFR" if diff < 0 else f"{int(diff)} SDCCH8"
                for diff in df_avg_topN['Diff']
            ]

            #######################################
            # Merge with TRXCHAN and do conversions
            #######################################
            df_comments = df_avg_topN[['Cell CI', 'Comments']]
            # Extract how many channels to add & type
            df_comments['Number To Add'] = df_comments['Comments'].str.extract(r'(\d+)').astype(float)
            df_comments['Type To Add']   = df_comments['Comments'].str.extract(r'([A-Za-z]+)')

            df_merged = pd.merge(self.df_trxchan, df_comments[['Cell CI', 'Number To Add', 'Type To Add']], on='Cell CI', how='left')
            df_merged['Converted Channel Type'] = df_merged['Channel Type']  # start with original

            # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            # Conversion for SDCCH
            # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            for cell_ci in df_merged['Cell CI'].unique():
                df_cell = df_merged[df_merged['Cell CI'] == cell_ci]
                # if 'Type To Add' is 'SDCCH8' ( or we check 'SDCCH' ) and > 0
                if df_cell['Type To Add'].iloc[0] == 'SDCCH' and not pd.isna(df_cell['Number To Add'].iloc[0]) and df_cell['Number To Add'].iloc[0] > 0:
                    channels_converted = 0
                    num_to_convert = int(df_cell['Number To Add'].iloc[0])
                    for idx, row in df_cell.iterrows():
                        if row['Converted Channel Type'] in ['TCHFR', 'TCHHR'] and row.get('Channel No.', 999) not in [6, 7]:
                            # also check there's more than 1 TCH in that TRX
                            df_trx = df_merged[
                                (df_merged['TRX No.'] == row['TRX No.']) &
                                (df_merged['Converted Channel Type'].isin(['TCHFR', 'TCHHR'])) &
                                (~df_merged['Channel No.'].isin([6, 7]))
                            ]
                            if len(df_trx) > 1:
                                df_merged.loc[idx, 'Converted Channel Type'] = 'SDCCH8'
                                channels_converted += 1
                                if channels_converted >= num_to_convert:
                                    break

            # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            # Conversion for PDCCH
            # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            for cell_ci in df_merged['Cell CI'].unique():
                df_cell = df_merged[df_merged['Cell CI'] == cell_ci]
                if df_cell['Type To Add'].iloc[0] == 'PDCCH' and not pd.isna(df_cell['Number To Add'].iloc[0]):
                    channels_converted = 0
                    num_to_convert = int(df_cell['Number To Add'].iloc[0])
                    # reverse order
                    for idx in df_cell.index[::-1]:
                        row = df_cell.loc[idx]
                        if row['Converted Channel Type'] in ['TCHFR', 'TCHHR']:
                            df_merged.loc[idx, 'Converted Channel Type'] = 'PDCCH'
                            channels_converted += 1
                            if channels_converted >= num_to_convert:
                                break

            # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            # Conversion for TCH
            # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            for cell_ci in df_merged['Cell CI'].unique():
                df_cell = df_merged[df_merged['Cell CI'] == cell_ci]
                if df_cell['Type To Add'].iloc[0] == 'TCHFR' and not pd.isna(df_cell['Number To Add'].iloc[0]):
                    channels_converted = 0
                    num_to_convert = int(df_cell['Number To Add'].iloc[0])
                    for idx, row in df_cell.iterrows():
                        if row['Converted Channel Type'] == 'SDCCH8':
                            if row.get('Is Main BCCH TRX') == 'YES':
                                non_main_bcch_left = df_cell[
                                    (df_cell['Converted Channel Type'] == 'SDCCH8') &
                                    (df_cell['Is Main BCCH TRX'] == 'NO')
                                ].shape[0] > 0
                                if non_main_bcch_left:
                                    continue
                            df_merged.loc[idx, 'Converted Channel Type'] = 'TCHFR'
                            channels_converted += 1
                            if channels_converted >= num_to_convert:
                                break

            # Post PDCH Channel Priority Type
            df_merged['Post PDCH Channel Priority Type'] = np.where(
                df_merged['Converted Channel Type'].isin(['PDCCH', 'TCHFR', 'TCHHR']),
                'EGPRSNORCH',
                ''
            )

            # Mark changed channels
            df_merged['Channel Type Changed'] = np.where(
                df_merged['Converted Channel Type'] == df_merged['Channel Type'],
                'No',
                'Yes'
            )

            df_merged_changed = df_merged[df_merged['Channel Type Changed'] == 'Yes'].copy()

            # Write out
            pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

            def remove_illegal_characters(value):
                if isinstance(value, str):
                    return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', value)
                return value

            book = Workbook()
            sheet = book.active

            # Dump data to sheet
            headers_done = False
            for row in dataframe_to_rows(df_merged_changed, index=False, header=True):
                cleaned_row = [remove_illegal_characters(cell) for cell in row]
                sheet.append(cleaned_row)
                # Color the header row once
                if not headers_done:
                    headers_done = True
                    header_cells = sheet[1]
                    for cell in header_cells:
                        if cell.value in ["Channel Type", "PDCH Channel Priority Type"]:
                            cell.fill = pink_fill
                        elif cell.value in ["Converted Channel Type", "Post PDCH Channel Priority Type"]:
                            cell.fill = green_fill

            output_filename = "Wap_NPM_Level1_SDCCH_dimensioning.xlsx"
            book.save(output_filename)

            self.process_status.config(text=f"Done! Output: {output_filename}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")
        finally:
            # Stop progress bar
            self.process_progress.stop()
            self.process_progress.config(mode='determinate', value=100)

##############################################
# The main entry point
##############################################
def main():
    root = tk.Tk()
    app = SDCCHDimensioningGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
